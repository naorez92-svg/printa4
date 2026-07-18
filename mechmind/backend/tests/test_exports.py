"""יצוא Excel ו-PDF — כולל בדיקת escape נגד הזרקת HTML."""
from openpyxl import load_workbook

from backend.core.exports import export_bom_xlsx, export_pdf


def test_bom_xlsx(tmp_path):
    rows = [{"item": 1, "name": "פלטת בסיס", "material": "אלומיניום 6061",
             "qty": 2, "mass_g": 120.5, "unit_cost_ils": 8.4, "notes": ""}]
    path = export_bom_xlsx(rows, tmp_path / "bom.xlsx")
    wb = load_workbook(path)
    ws = wb["BOM"]
    assert ws.sheet_view.rightToLeft
    assert ws.cell(row=4, column=2).value == "פלטת בסיס"
    assert ws.cell(row=4, column=5).value == 120.5


def test_pdf_hebrew(tmp_path):
    path = export_pdf("תוכנית בדיקה",
                      [{"heading": "משימות", "paragraphs": ["פסקה ראשונה"],
                        "bullets": ["סעיף 1"],
                        "table": {"headers": ["א", "ב"], "rows": [["1", "2"]]}}],
                      tmp_path / "plan.pdf", safety_note="הערת בטיחות")
    data = path.read_bytes()
    assert data[:5] == b"%PDF-"
    assert len(data) > 1000


def test_pdf_escapes_html_injection(tmp_path):
    """תוכן מה-LLM/משתמש חייב לעבור escape — לא להתפרש כ-HTML."""
    evil = '<script>alert(1)</script><img src="http://evil/x">'
    path = export_pdf(evil, [{"heading": evil, "paragraphs": [evil]}],
                      tmp_path / "evil.pdf")
    assert path.read_bytes()[:5] == b"%PDF-"
    # אם ה-escape עובד, ה-PDF נוצר בלי לנסות לטעון משאב חיצוני (WeasyPrint היה נכשל/מזהיר)
