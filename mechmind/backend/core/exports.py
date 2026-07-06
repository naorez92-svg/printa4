"""יצוא מסמכים: כתב כמויות ל-Excel (openpyxl) ומסמכי PDF בעברית (WeasyPrint)."""
from __future__ import annotations

import html
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

INK = "20184A"
BRAND = "F4A02C"


def export_bom_xlsx(rows: list[dict], out_path: str | Path, title: str = "כתב כמויות") -> Path:
    """rows: [{item, name, material, qty, mass_g, unit_cost_ils, notes}]"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = f"MechMind · {title}"
    cell.font = Font(bold=True, size=14, color=INK)
    cell.alignment = Alignment(horizontal="center")

    headers = ["פריט", "תיאור", "חומר", "כמות", "מסה (גרם)", "עלות ליח' (₪)", "הערות"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=INK)
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, start=4):
        values = [row.get("item", i - 3), row.get("name", ""), row.get("material", ""),
                  row.get("qty", 1), row.get("mass_g"), row.get("unit_cost_ils"),
                  row.get("notes", "")]
        for col, v in enumerate(values, start=1):
            ws.cell(row=i, column=col, value=v)

    for col, width in zip("ABCDEFG", (8, 34, 22, 8, 12, 14, 30)):
        ws.column_dimensions[col].width = width

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


_PDF_CSS = """
@page { size: A4; margin: 2cm; }
body { direction: rtl; font-family: 'DejaVu Sans', sans-serif; color: #20184A; font-size: 11pt; }
h1 { color: #20184A; border-bottom: 3px solid #F4A02C; padding-bottom: 6px; font-size: 18pt; }
h2 { color: #6C5CE7; font-size: 13pt; margin-top: 18px; }
table { width: 100%; border-collapse: collapse; margin: 8px 0; }
th { background: #20184A; color: white; padding: 6px 8px; text-align: right; font-size: 10pt; }
td { border-bottom: 1px solid #ddd; padding: 5px 8px; text-align: right; font-size: 10pt; }
.safety { background: #FFF3E5; border-right: 4px solid #F4A02C; padding: 10px 12px;
          margin: 14px 0; font-size: 9.5pt; }
.footer { margin-top: 24px; font-size: 8.5pt; color: #888; }
ul { margin: 4px 0; padding-right: 18px; }
"""


def export_pdf(title: str, sections: list[dict], out_path: str | Path,
               safety_note: str | None = None) -> Path:
    """sections: [{heading, paragraphs: [str], table: {headers: [], rows: [[]]}, bullets: [str]}]

    כל הטקסט עובר escape — אין הזרקת HTML מתוכן שנוצר על-ידי LLM או משתמש.
    """
    from weasyprint import HTML

    def _block_remote(url: str):
        # הגנה לעומק: התוכן כבר עובר escape, אבל חוסמים לחלוטין כל טעינת
        # משאב חיצוני (http/https/file) מתוך ה-HTML — רק data: URI מותר.
        if not url.startswith("data:"):
            raise ValueError(f"טעינת משאב חיצוני חסומה בייצוא PDF: {url[:40]}")
        from weasyprint.urls import default_url_fetcher
        return default_url_fetcher(url)

    parts = [f"<h1>{html.escape(title)}</h1>"]
    if safety_note:
        parts.append(f'<div class="safety">{html.escape(safety_note)}</div>')
    for sec in sections:
        if sec.get("heading"):
            parts.append(f"<h2>{html.escape(str(sec['heading']))}</h2>")
        for p in sec.get("paragraphs", []):
            parts.append(f"<p>{html.escape(str(p))}</p>")
        if sec.get("bullets"):
            items = "".join(f"<li>{html.escape(str(b))}</li>" for b in sec["bullets"])
            parts.append(f"<ul>{items}</ul>")
        if sec.get("table"):
            headers = "".join(f"<th>{html.escape(str(h))}</th>" for h in sec["table"].get("headers", []))
            rows = "".join(
                "<tr>" + "".join(f"<td>{html.escape(str(c))}</td>" for c in row) + "</tr>"
                for row in sec["table"].get("rows", [])
            )
            parts.append(f"<table><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>")
    parts.append('<div class="footer">הופק על-ידי MechMind · מהנדס-העל</div>')

    doc = f'<html lang="he" dir="rtl"><head><meta charset="utf-8"><style>{_PDF_CSS}</style></head><body>{"".join(parts)}</body></html>'
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    HTML(string=doc, url_fetcher=_block_remote).write_pdf(str(out))
    return out
